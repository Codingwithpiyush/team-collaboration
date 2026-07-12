import csv
import io
from decimal import Decimal
from django.db.models import Sum, Avg, Count
from django.http import HttpResponse
from django.utils import timezone
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from reportlab.lib.pagesizes import letter
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib import colors

# Import Models dynamically or directly to avoid import circular dependencies
from users.models import Department, EmployeeProfile, Category
from environmental.models import CarbonTransaction, EnvironmentalGoal, DepartmentEnvironmentalScore
from social.models import EmployeeParticipation, Training, DiversityMetric, DepartmentSocialScore
from governance.models import Policy, PolicyAcknowledgement, Audit, ComplianceIssue, DepartmentGovernanceScore
from gamification.models import Challenge, ChallengeParticipation, EmployeeXP, EmployeeBadge, RewardRedemption, EmployeeLeaderboard


class ReportEngine:

    @staticmethod
    def filter_queryset_by_common(queryset, filters, date_field='created_at'):
        """
        Helper to filter querysets by department, employee, status, and date range.
        """
        if 'department' in filters and filters['department']:
            # Handle relations dynamically
            if hasattr(queryset.model, 'department'):
                queryset = queryset.filter(department_id=filters['department'])
            elif hasattr(queryset.model, 'employee') and hasattr(EmployeeProfile, 'department'):
                queryset = queryset.filter(employee__department_id=filters['department'])

        if 'employee' in filters and filters['employee']:
            if hasattr(queryset.model, 'employee'):
                queryset = queryset.filter(employee_id=filters['employee'])
            elif queryset.model == EmployeeProfile:
                queryset = queryset.filter(id=filters['employee'])

        if 'status' in filters and filters['status']:
            if hasattr(queryset.model, 'status'):
                queryset = queryset.filter(status=filters['status'])

        # Date range filtering
        start_date = filters.get('start_date')
        end_date = filters.get('end_date')
        if start_date:
            queryset = queryset.filter(**{f"{date_field}__gte": start_date})
        if end_date:
            queryset = queryset.filter(**{f"{date_field}__lte": end_date})

        return queryset

    @classmethod
    def get_environmental_data(cls, filters):
        """
        Generates carbon ledger and environmental goal metrics.
        """
        txs = CarbonTransaction.objects.all().select_related('department', 'employee__user', 'emission_factor')
        txs = cls.filter_queryset_by_common(txs, filters, date_field='transaction_date')

        goals = EnvironmentalGoal.objects.all().select_related('department')
        if 'department' in filters and filters['department']:
            goals = goals.filter(department_id=filters['department'])
        if 'status' in filters and filters['status']:
            goals = goals.filter(status=filters['status'])

        total_co2 = txs.aggregate(total=Sum('calculated_co2'))['total'] or Decimal('0.0')

        summary = {
            'total_emissions_kg': float(total_co2),
            'transaction_count': txs.count(),
            'goals_count': goals.count(),
            'achieved_goals': goals.filter(status='achieved').count()
        }

        details = [
            {
                'date': tx.transaction_date.strftime('%Y-%m-%d'),
                'department': tx.department.name,
                'employee': tx.employee.user.get_full_name() if tx.employee else 'System',
                'activity': tx.activity_type,
                'quantity': float(tx.quantity),
                'unit': tx.unit,
                'co2_emitted_kg': float(tx.calculated_co2)
            }
            for tx in txs
        ]

        return {
            'title': 'Environmental Performance Report',
            'summary': summary,
            'headers': ['Date', 'Department', 'Employee', 'Activity', 'Quantity', 'Unit', 'CO2 Emitted (kg)'],
            'rows': [[d['date'], d['department'], d['employee'], d['activity'], d['quantity'], d['unit'], d['co2_emitted_kg']] for d in details],
            'details': details
        }

    @classmethod
    def get_social_data(cls, filters):
        """
        Generates CSR activities participation and training hour logs.
        """
        parts = EmployeeParticipation.objects.all().select_related('employee__user', 'activity__category')
        parts = cls.filter_queryset_by_common(parts, filters, date_field='joined_date')

        if 'category' in filters and filters['category']:
            parts = parts.filter(activity__category_id=filters['category'])

        trainings = Training.objects.all()
        if 'status' in filters and filters['status']:
            trainings = trainings.filter(status=filters['status'])

        total_points = parts.filter(status='approved').aggregate(total=Sum('points_awarded'))['total'] or 0

        summary = {
            'total_csr_points_awarded': total_points,
            'participation_count': parts.count(),
            'approved_participations': parts.filter(status='approved').count(),
            'training_sessions_count': trainings.count()
        }

        details = [
            {
                'date': p.joined_date.strftime('%Y-%m-%d') if p.joined_date else '',
                'employee': p.employee.user.get_full_name(),
                'activity': p.activity.title,
                'category': p.activity.category.name,
                'status': p.status,
                'points_awarded': p.points_awarded
            }
            for p in parts
        ]

        return {
            'title': 'Social Engagement & CSR Report',
            'summary': summary,
            'headers': ['Date Joined', 'Employee', 'CSR Activity', 'Category', 'Status', 'Points Awarded'],
            'rows': [[d['date'], d['employee'], d['activity'], d['category'], d['status'], d['points_awarded']] for d in details],
            'details': details
        }

    @classmethod
    def get_governance_data(cls, filters):
        """
        Generates compliance tracking and policy audits report.
        """
        issues = ComplianceIssue.objects.all().select_related('department', 'assigned_to__user')
        issues = cls.filter_queryset_by_common(issues, filters, date_field='created_at')

        audits = Audit.objects.all().select_related('department', 'auditor__user')
        audits = cls.filter_queryset_by_common(audits, filters, date_field='audit_date')

        avg_audit = audits.filter(status='published').aggregate(avg=Avg('score'))['avg'] or 0.0

        summary = {
            'total_compliance_issues': issues.count(),
            'open_issues': issues.filter(status__in=['open', 'in_progress']).count(),
            'overdue_issues': issues.filter(status='overdue').count(),
            'resolved_issues': issues.filter(status='resolved').count(),
            'audits_conducted': audits.count(),
            'average_audit_score': float(avg_audit)
        }

        details = [
            {
                'title': issue.title,
                'department': issue.department.name,
                'assigned_to': issue.assigned_to.user.get_full_name(),
                'severity': issue.severity,
                'status': issue.status,
                'due_date': issue.due_date.strftime('%Y-%m-%d')
            }
            for issue in issues
        ]

        return {
            'title': 'Compliance & Corporate Governance Report',
            'summary': summary,
            'headers': ['Issue Title', 'Department', 'Assigned To', 'Severity', 'Status', 'Due Date'],
            'rows': [[d['title'], d['department'], d['assigned_to'], d['severity'], d['status'], d['due_date']] for d in details],
            'details': details
        }

    @classmethod
    def get_department_data(cls, department_id, filters):
        """
        Generates environmental, social, governance, and user metrics for a specific department.
        """
        try:
            dept = Department.objects.get(pk=department_id)
        except Department.DoesNotExist:
            raise ValueError("Department not found.")

        # Scores
        env_score = DepartmentEnvironmentalScore.objects.filter(department=dept).first()
        soc_score = DepartmentSocialScore.objects.filter(department=dept).first()
        gov_score = DepartmentGovernanceScore.objects.filter(department=dept).first()

        # Employees
        employees = EmployeeProfile.objects.filter(department=dept, status='active')

        summary = {
            'department_name': dept.name,
            'department_code': dept.code,
            'employee_count': employees.count(),
            'environmental_score': float(env_score.score) if env_score else 100.0,
            'social_score': float(soc_score.score) if soc_score else 100.0,
            'governance_score': float(gov_score.score) if gov_score else 100.0,
        }

        details = [
            {
                'employee_id': emp.employee_id,
                'name': emp.user.get_full_name(),
                'designation': emp.designation or 'Staff',
                'joining_date': emp.joining_date.strftime('%Y-%m-%d') if emp.joining_date else 'N/A',
                'status': emp.status
            }
            for emp in employees
        ]

        return {
            'title': f"Department Report - {dept.name}",
            'summary': summary,
            'headers': ['Employee ID', 'Name', 'Designation', 'Joining Date', 'Status'],
            'rows': [[d['employee_id'], d['name'], d['designation'], d['joining_date'], d['status']] for d in details],
            'details': details
        }

    @classmethod
    def get_employee_data(cls, employee_id, filters):
        """
        Generates active metrics, badges, and participations list for an employee.
        """
        try:
            emp = EmployeeProfile.objects.select_related('user', 'department').get(pk=employee_id)
        except EmployeeProfile.DoesNotExist:
            raise ValueError("Employee profile not found.")

        # XP
        xp_profile = EmployeeXP.objects.filter(employee=emp).first()
        badges = EmployeeBadge.objects.filter(employee=emp).select_related('badge')
        redemptions = RewardRedemption.objects.filter(employee=emp, status='approved')

        summary = {
            'employee_name': emp.user.get_full_name(),
            'employee_id': emp.employee_id,
            'department': emp.department.name if emp.department else 'N/A',
            'xp': xp_profile.xp if xp_profile else 0,
            'level': xp_profile.level if xp_profile else 1,
            'badges_unlocked': badges.count(),
            'rewards_redeemed': redemptions.count()
        }

        details = [
            {
                'badge_name': eb.badge.name,
                'description': eb.badge.description,
                'unlocked_at': eb.unlocked_at.strftime('%Y-%m-%d')
            }
            for eb in badges
        ]

        return {
            'title': f"Employee Performance Report - {emp.user.get_full_name()}",
            'summary': summary,
            'headers': ['Badge Name', 'Description', 'Unlocked Date'],
            'rows': [[d['badge_name'], d['description'], d['unlocked_at']] for d in details],
            'details': details
        }

    @classmethod
    def get_esg_summary(cls, filters):
        """
        Aggregates environmental, social, and governance indicators at corporate scale.
        """
        departments = Department.objects.filter(status='active')
        total_co2 = CarbonTransaction.objects.aggregate(total=Sum('calculated_co2'))['total'] or Decimal('0.0')

        avg_env = DepartmentEnvironmentalScore.objects.aggregate(avg=Avg('score'))['avg'] or 100.0
        avg_soc = DepartmentSocialScore.objects.aggregate(avg=Avg('score'))['avg'] or 100.0
        avg_gov = DepartmentGovernanceScore.objects.aggregate(avg=Avg('score'))['avg'] or 100.0

        summary = {
            'active_departments': departments.count(),
            'total_carbon_emissions_kg': float(total_co2),
            'average_environmental_score': float(avg_env),
            'average_social_score': float(avg_soc),
            'average_governance_score': float(avg_gov),
            'overall_esg_index': float((avg_env + avg_soc + avg_gov) / Decimal('3.0'))
        }

        details = [
            {
                'department': dept.name,
                'env_score': float(DepartmentEnvironmentalScore.objects.filter(department=dept).first().score) if DepartmentEnvironmentalScore.objects.filter(department=dept).exists() else 100.0,
                'social_score': float(DepartmentSocialScore.objects.filter(department=dept).first().score) if DepartmentSocialScore.objects.filter(department=dept).exists() else 100.0,
                'gov_score': float(DepartmentGovernanceScore.objects.filter(department=dept).first().score) if DepartmentGovernanceScore.objects.filter(department=dept).exists() else 100.0,
            }
            for dept in departments
        ]

        return {
            'title': 'Corporate ESG Score Card Summary',
            'summary': summary,
            'headers': ['Department', 'Environmental Score', 'Social Score', 'Governance Score'],
            'rows': [[d['department'], d['env_score'], d['social_score'], d['gov_score']] for d in details],
            'details': details
        }

    @classmethod
    def get_custom_report(cls, filters):
        """
        Filters and builds reports based on custom selections (Module, Department, Employee, Date range, Status).
        """
        module = filters.get('module', 'environmental')

        if module == 'environmental':
            txs = CarbonTransaction.objects.all().select_related('department', 'employee__user', 'emission_factor')
            txs = cls.filter_queryset_by_common(txs, filters, date_field='transaction_date')
            headers = ['Date', 'Department', 'Activity', 'Qty', 'Unit', 'CO2 (kg)']
            rows = [[tx.transaction_date.strftime('%Y-%m-%d'), tx.department.name, tx.activity_type, float(tx.quantity), tx.unit, float(tx.calculated_co2)] for tx in txs]
            details = [{'date': tx.transaction_date.strftime('%Y-%m-%d'), 'department': tx.department.name, 'activity': tx.activity_type, 'quantity': float(tx.quantity), 'unit': tx.unit, 'co2': float(tx.calculated_co2)} for tx in txs]

        elif module == 'social':
            parts = EmployeeParticipation.objects.all().select_related('employee__user', 'activity')
            parts = cls.filter_queryset_by_common(parts, filters, date_field='joined_date')
            headers = ['Date Joined', 'Employee', 'Activity', 'Status', 'Points']
            rows = [[p.joined_date.strftime('%Y-%m-%d') if p.joined_date else '', p.employee.user.get_full_name(), p.activity.title, p.status, p.points_awarded] for p in parts]
            details = [{'date': p.joined_date.strftime('%Y-%m-%d') if p.joined_date else '', 'employee': p.employee.user.get_full_name(), 'activity': p.activity.title, 'status': p.status, 'points': p.points_awarded} for p in parts]

        elif module == 'governance':
            issues = ComplianceIssue.objects.all().select_related('department', 'assigned_to__user')
            issues = cls.filter_queryset_by_common(issues, filters, date_field='created_at')
            headers = ['Issue Title', 'Department', 'Assigned To', 'Severity', 'Status', 'Due Date']
            rows = [[i.title, i.department.name, i.assigned_to.user.get_full_name(), i.severity, i.status, i.due_date.strftime('%Y-%m-%d')] for i in issues]
            details = [{'title': i.title, 'department': i.department.name, 'assigned_to': i.assigned_to.user.get_full_name(), 'severity': i.severity, 'status': i.status, 'due_date': i.due_date.strftime('%Y-%m-%d')} for i in issues]

        elif module == 'gamification':
            parts = ChallengeParticipation.objects.all().select_related('challenge', 'employee__user')
            parts = cls.filter_queryset_by_common(parts, filters, date_field='joined_date')
            headers = ['Date Joined', 'Employee', 'Challenge', 'Status']
            rows = [[p.joined_date.strftime('%Y-%m-%d') if p.joined_date else '', p.employee.user.get_full_name(), p.challenge.title, p.status] for p in parts]
            details = [{'date': p.joined_date.strftime('%Y-%m-%d') if p.joined_date else '', 'employee': p.employee.user.get_full_name(), 'challenge': p.challenge.title, 'status': p.status} for p in parts]

        else:
            raise ValueError("Unsupported Module selected.")

        return {
            'title': f"Custom {module.capitalize()} Builder Report",
            'summary': {'record_count': len(rows)},
            'headers': headers,
            'rows': rows,
            'details': details
        }

    # --- Export Engine ---

    @staticmethod
    def export_as_csv(title, headers, rows):
        """
        Outputs response data stream formatted as standard CSV comma separated values.
        """
        buffer = io.StringIO()
        writer = csv.writer(buffer)
        writer.writerow([title])
        writer.writerow([])
        writer.writerow(headers)
        for row in rows:
            writer.writerow(row)
        return buffer.getvalue()

    @staticmethod
    def export_as_excel(title, headers, rows, summary=None):
        """
        Creates binary workbook content structured with headings and cells using openpyxl.
        """
        wb = Workbook()
        ws = wb.active
        ws.title = "Report Details"

        # Styling
        title_font = Font(name='Arial', size=16, bold=True, color='FFFFFF')
        header_font = Font(name='Arial', size=11, bold=True, color='FFFFFF')
        bold_font = Font(name='Arial', size=11, bold=True)
        normal_font = Font(name='Arial', size=10)

        title_fill = PatternFill(start_color="1F2937", end_color="1F2937", fill_type="solid")
        header_fill = PatternFill(start_color="374151", end_color="374151", fill_type="solid")
        summary_fill = PatternFill(start_color="F3F4F6", end_color="F3F4F6", fill_type="solid")

        # Set Title
        ws.merge_cells('A1:G1')
        ws['A1'] = title
        ws['A1'].font = title_font
        ws['A1'].fill = title_fill
        ws['A1'].alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[1].height = 40

        curr_row = 3
        # Set Summary Info
        if summary:
            ws.cell(row=curr_row, column=1, value="KPI Metrics Summary").font = bold_font
            curr_row += 1
            for k, v in summary.items():
                ws.cell(row=curr_row, column=1, value=k.replace('_', ' ').capitalize()).font = bold_font
                ws.cell(row=curr_row, column=1).fill = summary_fill
                ws.cell(row=curr_row, column=2, value=v).font = normal_font
                curr_row += 1
            curr_row += 1

        # Headers
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=curr_row, column=col_num, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[curr_row].height = 25
        curr_row += 1

        # Rows
        for r in rows:
            for col_num, val in enumerate(r, 1):
                cell = ws.cell(row=curr_row, column=col_num, value=val)
                cell.font = normal_font
            curr_row += 1

        # Auto-fit column widths
        for col in ws.columns:
            max_len = max(len(str(cell.value or '')) for cell in col)
            col_letter = col[0].column_letter
            ws.column_dimensions[col_letter].width = max(max_len + 3, 12)

        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        return buffer.getvalue()

    @staticmethod
    def export_as_pdf(title, headers, rows, summary=None):
        """
        Uses reportlab templates to generate clean formatting.
        """
        buffer = io.BytesIO()
        doc = SimpleDocTemplate(
            buffer,
            pagesize=letter,
            rightMargin=36,
            leftMargin=36,
            topMargin=36,
            bottomMargin=36
        )
        styles = getSampleStyleSheet()

        # Custom Styles
        title_style = ParagraphStyle(
            'ReportTitle',
            parent=styles['Heading1'],
            fontSize=20,
            textColor=colors.HexColor('#1f2937'),
            spaceAfter=15
        )
        section_style = ParagraphStyle(
            'ReportSection',
            parent=styles['Heading2'],
            fontSize=13,
            textColor=colors.HexColor('#374151'),
            spaceAfter=8,
            spaceBefore=12
        )
        normal_style = styles['Normal']

        elements = []
        elements.append(Paragraph(title, title_style))
        elements.append(Spacer(1, 10))

        # Add Summary Metrics
        if summary:
            elements.append(Paragraph("<b>Summary KPIs</b>", section_style))
            summary_data = []
            for k, v in summary.items():
                label = k.replace('_', ' ').capitalize()
                summary_data.append([
                    Paragraph(f"<b>{label}</b>", normal_style),
                    Paragraph(str(v), normal_style)
                ])
            st = Table(summary_data, colWidths=[200, 200])
            st.setStyle(TableStyle([
                ('BACKGROUND', (0,0), (-1,-1), colors.HexColor('#f9fafb')),
                ('GRID', (0,0), (-1,-1), 0.5, colors.HexColor('#e5e7eb')),
                ('PADDING', (0,0), (-1,-1), 5),
            ]))
            elements.append(st)
            elements.append(Spacer(1, 15))

        # Add Main Table
        elements.append(Paragraph("<b>Detailed Logs</b>", section_style))
        table_data = []
        table_data.append([Paragraph(f"<b>{h}</b>", normal_style) for h in headers])
        for r in rows:
            table_data.append([Paragraph(str(cell), normal_style) for cell in r])

        t = Table(table_data)
        t.setStyle(TableStyle([
            ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#f3f4f6')),
            ('ALIGN', (0,0), (-1,-1), 'LEFT'),
            ('VALIGN', (0,0), (-1,-1), 'TOP'),
            ('GRID', (0,0), (-1,-1), 0.5, colors.HexColor('#d1d5db')),
            ('TOPPADDING', (0,0), (-1,-1), 6),
            ('BOTTOMPADDING', (0,0), (-1,-1), 6),
        ]))
        elements.append(t)

        doc.build(elements)
        buffer.seek(0)
        return buffer.getvalue()
