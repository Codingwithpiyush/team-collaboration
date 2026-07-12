from rest_framework import viewsets, status, filters
from rest_framework.decorators import action
from rest_framework.response import Response
from django_filters.rest_framework import DjangoFilterBackend
from rest_framework.pagination import PageNumberPagination
from rest_framework.permissions import AllowAny

from .models import Department, Category, EmployeeProfile, OrganizationSettings, NotificationSettings
from .serializers import (
    DepartmentSerializer,
    CategorySerializer,
    EmployeeProfileSerializer,
    OrganizationSettingsSerializer,
    NotificationSettingsSerializer
)
from .services import DepartmentService, OrganizationSettingsService


class StandardResultsSetPagination(PageNumberPagination):
    """
    Standard pagination for ESG platforms.
    """
    page_size = 10
    page_size_query_param = 'page_size'
    max_page_size = 100


class DepartmentViewSet(viewsets.ModelViewSet):
    """
    ViewSet for Department CRUD, hierarchy representation, and dropdown selection.
    """
    queryset = Department.objects.all().select_related('parent_department', 'head__user')
    serializer_class = DepartmentSerializer
    pagination_class = StandardResultsSetPagination
    permission_classes = [AllowAny]
    filter_backends = [DjangoFilterBackend, filters.SearchFilter, filters.OrderingFilter]
    filterset_fields = ['status', 'parent_department']
    search_fields = ['name', 'code']
    ordering_fields = ['name', 'code', 'employee_count']
    ordering = ['name']

    @action(detail=False, methods=['get'], url_path='hierarchy')
    def hierarchy(self, request):
        """
        Custom endpoint to fetch tree hierarchy of active departments.
        """
        hierarchy_tree = DepartmentService.get_hierarchy()
        return Response(hierarchy_tree, status=status.HTTP_200_OK)

    @action(detail=False, methods=['get'], url_path='dropdown')
    def dropdown(self, request):
        """
        Custom lightweight endpoint for fetching active departments for dropdowns.
        """
        active_depts = self.queryset.filter(status='active').values('id', 'name', 'code')
        return Response(active_depts, status=status.HTTP_200_OK)

    @action(detail=False, methods=['get'], url_path='statistics')
    def statistics(self, request):
        """
        Gets department and employee count summary metrics.
        """
        total_departments = Department.objects.count()
        active_departments = Department.objects.filter(status='active').count()
        inactive_departments = Department.objects.filter(status='inactive').count()
        total_employees = EmployeeProfile.objects.count()

        return Response({
            'total_departments': total_departments,
            'active_departments': active_departments,
            'inactive_departments': inactive_departments,
            'total_employees': total_employees
        }, status=status.HTTP_200_OK)

    @action(detail=False, methods=['post'], url_path='import')
    def bulk_import(self, request):
        """
        Imports departments from Excel or CSV files.
        """
        import_file = request.FILES.get('file')
        if not import_file:
            return Response({"detail": "No file uploaded. Please upload a file with key 'file'."}, status=status.HTTP_400_BAD_REQUEST)

        file_name = import_file.name.lower()
        departments_created = 0
        errors = []

        import io

        if file_name.endswith('.csv'):
            try:
                decoded_file = import_file.read().decode('utf-8')
                io_string = io.StringIO(decoded_file)
                import csv
                reader = csv.DictReader(io_string)
                for row_idx, row in enumerate(reader, start=1):
                    name = row.get('name') or row.get('Name')
                    code = row.get('code') or row.get('Code')
                    status_val = row.get('status', 'active').lower() or row.get('Status', 'active').lower()
                    parent_id = row.get('parent_department') or row.get('Parent Department')
                    head_id = row.get('head') or row.get('Head')

                    if not name or not code:
                        errors.append(f"Row {row_idx}: 'name' and 'code' are required.")
                        continue

                    code = code.upper()

                    parent = None
                    if parent_id:
                        try:
                            parent = Department.objects.get(pk=parent_id)
                        except Department.DoesNotExist:
                            errors.append(f"Row {row_idx}: Parent department with ID {parent_id} does not exist.")
                            continue

                    head = None
                    if head_id:
                        try:
                            head = EmployeeProfile.objects.get(pk=head_id)
                        except EmployeeProfile.DoesNotExist:
                            errors.append(f"Row {row_idx}: Head employee with ID {head_id} does not exist.")
                            continue

                    dept, created = Department.objects.update_or_create(
                        code=code,
                        defaults={
                            'name': name,
                            'status': status_val if status_val in ['active', 'inactive'] else 'active',
                            'parent_department': parent,
                            'head': head
                        }
                    )
                    if created:
                        departments_created += 1

            except Exception as e:
                return Response({"detail": f"Error parsing CSV: {str(e)}"}, status=status.HTTP_400_BAD_REQUEST)

        elif file_name.endswith('.xlsx') or file_name.endswith('.xls'):
            try:
                from openpyxl import load_workbook
                wb = load_workbook(import_file)
                ws = wb.active

                headers = [cell.value for cell in ws[1]]
                header_map = {h.lower() if h else '': idx for idx, h in enumerate(headers)}

                for row_idx in range(2, ws.max_row + 1):
                    row_cells = [ws.cell(row=row_idx, column=col_idx).value for col_idx in range(1, len(headers) + 1)]
                    if not any(row_cells):
                        continue

                    def get_cell_value(field_name):
                        idx = header_map.get(field_name.lower())
                        if idx is not None and idx < len(row_cells):
                            return row_cells[idx]
                        return None

                    name = get_cell_value('name') or get_cell_value('Name')
                    code = get_cell_value('code') or get_cell_value('Code')
                    status_val = str(get_cell_value('status') or get_cell_value('Status') or 'active').lower()
                    parent_id = get_cell_value('parent_department') or get_cell_value('Parent Department')
                    head_id = get_cell_value('head') or get_cell_value('Head')

                    if not name or not code:
                        errors.append(f"Row {row_idx}: 'name' and 'code' are required.")
                        continue

                    code = str(code).upper()

                    parent = None
                    if parent_id:
                        try:
                            parent = Department.objects.get(pk=int(parent_id))
                        except (Department.DoesNotExist, ValueError):
                            errors.append(f"Row {row_idx}: Parent department with ID {parent_id} does not exist.")
                            continue

                    head = None
                    if head_id:
                        try:
                            head = EmployeeProfile.objects.get(pk=int(head_id))
                        except (EmployeeProfile.DoesNotExist, ValueError):
                            errors.append(f"Row {row_idx}: Head employee with ID {head_id} does not exist.")
                            continue

                    dept, created = Department.objects.update_or_create(
                        code=code,
                        defaults={
                            'name': str(name),
                            'status': status_val if status_val in ['active', 'inactive'] else 'active',
                            'parent_department': parent,
                            'head': head
                        }
                    )
                    if created:
                        departments_created += 1

            except Exception as e:
                return Response({"detail": f"Error parsing Excel: {str(e)}"}, status=status.HTTP_400_BAD_REQUEST)

        else:
            return Response({"detail": "Unsupported file format. Please upload a CSV or Excel file."}, status=status.HTTP_400_BAD_REQUEST)

        return Response({
            "departments_created": departments_created,
            "errors": errors
        }, status=status.HTTP_201_CREATED if departments_created > 0 else status.HTTP_200_OK)

    @action(detail=False, methods=['get'], url_path='export')
    def bulk_export(self, request):
        """
        Exports active departments in PDF, Excel, or CSV formats.
        """
        queryset = self.queryset.all()

        status_param = request.query_params.get('status')
        if status_param:
            queryset = queryset.filter(status=status_param)

        export_format = request.query_params.get('export', 'csv').lower()

        title = "Departments Export Report"
        headers = ["ID", "Name", "Code", "Head", "Parent Department", "Status", "Employee Count"]

        rows = []
        for dept in queryset:
            head_name = dept.head.user.get_full_name() if dept.head else ''
            parent_name = dept.parent_department.name if dept.parent_department else ''
            rows.append([
                str(dept.id),
                dept.name,
                dept.code,
                head_name,
                parent_name,
                dept.get_status_display() if hasattr(dept, 'get_status_display') else dept.status,
                str(dept.employee_count)
            ])

        summary = {
            "Total Departments": len(rows),
            "Active Departments": queryset.filter(status='active').count(),
            "Inactive Departments": queryset.filter(status='inactive').count()
        }

        from reports.services import ReportEngine
        from django.http import HttpResponse

        if export_format == 'csv':
            csv_data = ReportEngine.export_as_csv(title, headers, rows)
            response = HttpResponse(csv_data, content_type='text/csv')
            response['Content-Disposition'] = 'attachment; filename="departments.csv"'
            return response

        elif export_format == 'excel':
            excel_data = ReportEngine.export_as_excel(title, headers, rows, summary)
            response = HttpResponse(excel_data, content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
            response['Content-Disposition'] = 'attachment; filename="departments.xlsx"'
            return response

        elif export_format == 'pdf':
            pdf_data = ReportEngine.export_as_pdf(title, headers, rows, summary)
            response = HttpResponse(pdf_data, content_type='application/pdf')
            response['Content-Disposition'] = 'attachment; filename="departments.pdf"'
            return response

        return Response({"detail": "Invalid export format"}, status=status.HTTP_400_BAD_REQUEST)


class CategoryViewSet(viewsets.ModelViewSet):
    """
    ViewSet for Category CRUD and dropdown selection.
    """
    queryset = Category.objects.all()
    serializer_class = CategorySerializer
    pagination_class = StandardResultsSetPagination
    permission_classes = [AllowAny]
    filter_backends = [DjangoFilterBackend, filters.SearchFilter, filters.OrderingFilter]
    filterset_fields = ['status', 'type']
    search_fields = ['name', 'description']
    ordering_fields = ['name']
    ordering = ['name']

    @action(detail=False, methods=['get'], url_path='dropdown')
    def dropdown(self, request):
        """
        Custom lightweight endpoint for fetching active categories for dropdowns.
        """
        active_cats = self.queryset.filter(status='active').values('id', 'name', 'type')
        return Response(active_cats, status=status.HTTP_200_OK)


class EmployeeProfileViewSet(viewsets.ModelViewSet):
    """
    ViewSet for EmployeeProfile CRUD.
    """
    queryset = EmployeeProfile.objects.all().select_related('user', 'department')
    serializer_class = EmployeeProfileSerializer
    pagination_class = StandardResultsSetPagination
    permission_classes = [AllowAny]
    filter_backends = [DjangoFilterBackend, filters.SearchFilter, filters.OrderingFilter]
    filterset_fields = ['status', 'department', 'joining_date']
    search_fields = ['employee_id', 'user__username', 'user__first_name', 'user__last_name', 'designation']
    ordering_fields = ['employee_id', 'joining_date']
    ordering = ['employee_id']

    @action(detail=False, methods=['get'], url_path='dropdown')
    def dropdown(self, request):
        """
        Custom lightweight endpoint for fetching active employees for dropdowns.
        """
        active_employees = self.queryset.filter(status='active').values(
            'id', 'employee_id', 'designation'
        )
        # Add full name manually to response mapping
        response_data = []
        for emp in self.queryset.filter(status='active'):
            response_data.append({
                'id': emp.id,
                'employee_id': emp.employee_id,
                'name': emp.user.get_full_name() or emp.user.username,
                'designation': emp.designation
            })
        return Response(response_data, status=status.HTTP_200_OK)


class OrganizationSettingsViewSet(viewsets.ModelViewSet):
    """
    ViewSet for accessing and updating the global Organization Settings singleton.
    """
    queryset = OrganizationSettings.objects.all()
    serializer_class = OrganizationSettingsSerializer
    permission_classes = [AllowAny]

    def get_object(self):
        return OrganizationSettingsService.get_settings()

    def list(self, request, *args, **kwargs):
        instance = self.get_object()
        serializer = self.get_serializer(instance)
        return Response(serializer.data, status=status.HTTP_200_OK)

    def create(self, request, *args, **kwargs):
        # Create requests are converted to update on singleton
        instance = self.get_object()
        serializer = self.get_serializer(instance, data=request.data, partial=True)
        serializer.is_valid(raise_exception=True)
        self.perform_update(serializer)
        return Response(serializer.data, status=status.HTTP_200_OK)


class NotificationSettingsViewSet(viewsets.ModelViewSet):
    """
    ViewSet for managing individual notification preferences.
    """
    queryset = NotificationSettings.objects.all().select_related('employee__user')
    serializer_class = NotificationSettingsSerializer
    pagination_class = StandardResultsSetPagination
    permission_classes = [AllowAny]
    filter_backends = [DjangoFilterBackend, filters.OrderingFilter]
    filterset_fields = [
        'employee', 'email_notifications', 'badge_notifications',
        'csr_notifications', 'challenge_notifications', 'policy_notifications',
        'compliance_notifications'
    ]
    ordering_fields = ['id']
    ordering = ['id']


class SettingsDashboardViewSet(viewsets.ViewSet):
    """
    ViewSet for the corporate Settings and Configurations Dashboard.
    """
    permission_classes = [AllowAny]

    @action(detail=False, methods=['get'], url_path='dashboard')
    def dashboard(self, request):
        org_settings = OrganizationSettingsService.get_settings()
        org_serializer = OrganizationSettingsSerializer(org_settings)

        notif_settings = NotificationSettings.objects.all()
        notif_serializer = NotificationSettingsSerializer(notif_settings, many=True)

        total_departments = Department.objects.count()
        active_departments = Department.objects.filter(status='active').count()
        inactive_departments = Department.objects.filter(status='inactive').count()
        total_employees = EmployeeProfile.objects.count()

        dept_stats = {
            'total_departments': total_departments,
            'active_departments': active_departments,
            'inactive_departments': inactive_departments,
            'total_employees': total_employees
        }

        category_count = Category.objects.count()

        return Response({
            'organization_settings': org_serializer.data,
            'notification_settings': notif_serializer.data,
            'department_statistics': dept_stats,
            'category_count': category_count
        }, status=status.HTTP_200_OK)
